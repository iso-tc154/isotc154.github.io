#!/usr/bin/env python3
"""
Import plenary meeting list from an ISO/CS export (xlsx) into data/meetings.yml.

Usage:
    python3 scripts/import-meetings-xlsx.py /path/to/Meetings.xlsx

The xlsx is treated as the canonical list of TC 154 plenary meetings.
We keep:
  - ordinal (the plenary number, e.g. 45)
  - iso_meeting_id (the ISO/CS platform ID, e.g. 182492)
  - type (Face-to-Face / Hybrid / Virtual)
  - status (Closed / Cancelled / Registration open / etc.)
  - start_date, end_date  (LOCAL date+time, with tz abbrev preserved as-is)
  - country, city, address
  - virtual_address (Zoom/Webex link or label)
  - participants (attended participant count)
  - parent_iso_meeting_id (when one plenary was split into multiple sub-sessions)
  - cancellation_comment, reschedule_note

The YAML emitted here is the canonical meeting list. Rich per-plenary detail
(venue with map, schedule, secretariat, accommodation, practical info, etc.)
stays in _data/events/plenary-meeting-{n}.yml and is merged at build time.

Run this whenever ISO/CS sends an updated meeting export.
"""

from __future__ import annotations

import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

try:
    import yaml
except ImportError:
    sys.exit("PyYAML is required: pip install pyyaml")


REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = REPO_ROOT / "data" / "meetings.yml"


TYPE_NORMALISE = {
    "Face-to-Face": "face-to-face",
    "Hybrid": "hybrid",
    "Virtual": "virtual",
}

STATUS_NORMALISE = {
    "Closed": "closed",
    "Cancelled": "cancelled",
    "Registration open": "registration-open",
    "Confirmed": "confirmed",
    "Tentative": "tentative",
    "Scheduled": "scheduled",
}


def clean(s):
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None


def parse_local_datetime(raw):
    """
    Preserve the local date+time WITH the original timezone abbreviation.
    Examples from the export:
      '31 Aug 2026  00:00 CEST'
      '04 Oct 1979  09:00 HKST'
      '02 Sep 2021  00:00 Z'        (Zulu = UTC, written without offset)
    Normalise spacing to a single space; leave tz abbrev as-is.
    """
    if not raw:
        return None
    s = str(raw).strip()
    # collapse runs of whitespace
    s = re.sub(r"\s+", " ", s)
    return s


def iso_meeting_url(meeting_id):
    if not meeting_id:
        return None
    return f"https://committee.iso.org/meetings/{int(meeting_id)}"


def main():
    if len(sys.argv) < 2:
        sys.exit(f"Usage: {sys.argv[0]} <Meetings.xlsx>")

    xlsx_path = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx_path.exists():
        sys.exit(f"Not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Meetings"] if "Meetings" in wb.sheetnames else wb[wb.sheetnames[0]]

    header = [clean(c) for c in next(ws.iter_rows(values_only=True))]
    col = {name: i for i, name in enumerate(header) if name}

    required = [
        "Number", "Type", "Status", "Start date", "End date",
        "Country", "City", "meetingId",
    ]
    missing = [c for c in required if c not in col]
    if missing:
        sys.exit(f"Missing required columns: {missing}. Found: {header}")

    meetings = []
    for row in ws.iter_rows(values_only=True):
        # skip header
        if row[col["Number"]] is None:
            continue

        ordinal = row[col["Number"]]
        try:
            ordinal_int = int(ordinal)
        except (TypeError, ValueError):
            continue

        m = {
            "ordinal": ordinal_int,
            "iso_meeting_id": int(row[col["meetingId"]]) if row[col["meetingId"]] else None,
            "type": TYPE_NORMALISE.get(clean(row[col["Type"]]), clean(row[col["Type"]])),
            "status": STATUS_NORMALISE.get(clean(row[col["Status"]]), clean(row[col["Status"]])),
            "start_date": parse_local_datetime(row[col["Start date"]]),
            "end_date": parse_local_datetime(row[col["End date"]]),
            "country": clean(row[col["Country"]]) if "Country" in col else None,
            "city": clean(row[col["City"]]) if "City" in col else None,
            "address": clean(row[col["Address"]]) if "Address" in col else None,
            "virtual_address": clean(row[col["Virtual address"]]) if "Virtual address" in col else None,
            "participants": int(row[col["Attended participant"]]) if "Attended participant" in col and row[col["Attended participant"]] else None,
            "parent_iso_meeting_id": int(row[col["parent"]]) if "parent" in col and row[col["parent"]] else None,
            "reschedule_note": clean(row[col["Plan for the reschedule"]]) if "Plan for the reschedule" in col else None,
            "reschedule_timeframe": clean(row[col["Meeting reschedule timeframe"]]) if "Meeting reschedule timeframe" in col else None,
            "cancellation_comment": clean(row[col["Cancellation comment"]]) if "Cancellation comment" in col else None,
        }

        if m["iso_meeting_id"]:
            m["iso_meeting_url"] = iso_meeting_url(m["iso_meeting_id"])

        # The export includes English + French titles; we don't need them in the
        # canonical list (the committee title is the same for every plenary).

        meetings.append(m)

    # Sort newest first for readability
    meetings.sort(key=lambda x: (-x["ordinal"], x["start_date"] or ""))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8") as f:
        f.write("# Canonical list of ISO/TC 154 plenary meetings.\n")
        f.write("# Source: ISO/CS meeting export (Meetings.xlsx).\n")
        f.write("# Regenerate with: python3 scripts/import-meetings-xlsx.py <Meetings.xlsx>\n")
        f.write(f"# {len(meetings)} meetings, ordinals {meetings[-1]['ordinal']}–{meetings[0]['ordinal']}.\n\n")
        yaml.dump(meetings, f, sort_keys=False, allow_unicode=True, width=100)

    print(f"Wrote {OUTPUT.relative_to(REPO_ROOT)} ({len(meetings)} meetings)")


if __name__ == "__main__":
    main()
